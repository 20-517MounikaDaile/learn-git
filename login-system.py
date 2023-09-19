from tkinter import *
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib
from datetime import date
from tkinter.ttk import Combobox
from tkinter import filedialog
import os
from PIL import Image,ImageTk
from tkinter import messagebox




background="#06283D"
framebg="#EDEDED"
framefg="#06283D"
gender="Male"

root=Tk()
root.title("Student registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)
file=pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of Registration"
    sheet['G1']="Religion"
    sheet['H1']="Skill"
    sheet['I1']="Father Name"
    sheet['J1']="Mother Name"
    sheet['K1']="Father's Occupation"
    sheet['L1']="Mother's Occupation"
    file.save('Student_data.xlsx')
#Exit
def Exit():
    root.destroy()



def search():
    text=Search.get()

    Clear()
    saveButton.config(state='disable')
    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==int(text):
            name=row[0]

            print(str(name))
    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid Registration number")
        
                             
#gender
def selection():
    value=radio.get()
    if value==1:
        gender="Male"
        print(gender)
       
    else:
        gender="Female"
        print(gender)

def showimage():
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),title="Select image file",filetype=(("JPG File",".jpg"),("PNG File",".png"),("All files","*.txt")))
    img=(Image.open(filename))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2
def registration_no():
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row=sheet.max_row
   


    max_row_value=sheet.cell(row=row,column=1).value
    print(max_row_value)
    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")
def Clear():
    name.set('')
    dob.set('')
    religion.set('')
    skills.set('')
    fname.set('')
    mname.set('')
    fo.set('')
    mo.set('')
    Class.set("Select Class")

    registration_no()

    saveButton.config(state='normal')
    img1=PhotoImage(file="Student Registration System-images\Images/upload photo.png")
    lbl.config(image=img1)
    lbl.image=img1

    img=""
def Save():
    R1=Registration.get()
    N1=name.get()
    C1=Class.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender!")
    D2=dob.get()
    D1=Date.get()
    Re1=religion.get()
    S1=skills.get()
    fathername=fname.get()
    mothername=mname.get()
    F1=fo.get()
    M1=mo.get()

    if N1=="" or C1=="Select Class" or D2=="" or Re1=="" or S1=="" or fathername=="" or mothername=="" or F1=="" or M1=="" or S1=="":
        messagebox.showerror('error',"Few Data is missing!")
    else:
        file=openpyxl.load_workbook('Student_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=C1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=D2)
        sheet.cell(column=6,row=sheet.max_row,value=D1)
        sheet.cell(column=7,row=sheet.max_row,value=Re1)
        sheet.cell(column=8,row=sheet.max_row,value=S1)
        sheet.cell(column=9,row=sheet.max_row,value=fathername)
        sheet.cell(column=10,row=sheet.max_row,value=mothername)
        sheet.cell(column=11,row=sheet.max_row,value=F1)
        sheet.cell(column=12,row=sheet.max_row,value=M1)

        file.save(r'Student_data.xlsx')
        try:
            img.save("Student Registration System-images\Student Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profile Picture is not available!!!!")
        messagebox.showinfo("info","Successfully data entered")
        Clear()
        registration_no()
        
        



        
        
    
        
Label(root,text="Email:parvatcomputertechnology@gmail.com",width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION",width=10,height=2,bg="#c36464",fg='#fff',font="arial 20 bold",anchor=CENTER).pack(side=TOP,fill=X)
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=860,y=70)
imageicon3=PhotoImage(file="Student Registration System-images\Images\search.png")
Srch=Button(root,text="Search",compound=LEFT,image=imageicon3,width=120,bg="#68ddfa",font="arial 13 bold",command=search)
Srch.place(x=1100,y=66)

imageicon4=PhotoImage(file="Student Registration System-images\Images\Layer 4.png")
Update_button=Button(root,image=imageicon4,bg="#c36464")
Update_button.place(x=110,y=66)
Label(root,text="Registration No: ",font="arial 13",bg=background,fg=framebg).place(x=30,y=150)
Label(root,text="Date: ",font="arial 13",bg=background,fg=framebg).place(x=500,y=150)


Registration=StringVar()
Date=StringVar()

reg_entry=Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=150)

registration_no()


today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)

Date.set(d1)


#Student details

obj=LabelFrame(root,text="Student Details",font=20,bd=1,width=900,bg=framebg,height=220,relief=GROOVE)
obj.place(x=30,y=200)
Label(obj,text="Full Name: ",font="arial 13",bg=framebg).place(x=30,y=50)
Label(obj,text="Date of Birth: ",font="arial 13",bg=framebg).place(x=30,y=100)
Label(obj,text="Gender: ",font="arial 13",bg=framebg).place(x=30,y=150)

Label(obj,text="Class: ",font="arial 13",bg=framebg).place(x=500,y=50)
Label(obj,text="Religion: ",font="arial 13",bg=framebg).place(x=500,y=100)
Label(obj,text="Skills: ",font="arial 13",bg=framebg).place(x=500,y=150)

name=StringVar()
name_entry=Entry(obj,textvariable=name,width=15,font="arial 10")
name_entry.place(x=165,y=50)

dob=StringVar()
dob_entry=Entry(obj,textvariable=dob,width=15,font="arial 10")
dob_entry.place(x=165,y=100)


radio=IntVar()
R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,command=selection)
R1.place(x=150,y=150)

R2=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,command=selection)
R2.place(x=200,y=150)


religion=StringVar()
religion_entry=Entry(obj,textvariable=religion,width=15,font="arial 10")
religion_entry.place(x=650,y=100)


skills=StringVar()
skills_entry=Entry(obj,textvariable=skills,width=15,font="arial 10")
skills_entry.place(x=650,y=150)

Class=Combobox(obj,values=["1","2","3","4","5","6","7","8","9","10","11","12"],font="Robot 10",width=17,state='r')
Class.place(x=630,y=50)
Class.set("Select Class")

#Parents details
obj2=LabelFrame(root,text="Parent Details",font=20,bd=1,width=900,bg=framebg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Father's Name: ",font="arial 13",bg=framebg).place(x=30,y=50)
Label(obj2,text="Occupation: ",font="arial 13",bg=framebg).place(x=30,y=100)

fname=StringVar()
fname_entry=Entry(obj2,textvariable=fname,width=15,font="arial 10")
fname_entry.place(x=165,y=50)
fo=StringVar()
fo_entry=Entry(obj2,textvariable=fo,width=15,font="arial 10")
fo_entry.place(x=165,y=100)


Label(obj2,text="Mother's Name: ",font="arial 13",bg=framebg).place(x=500,y=50)
Label(obj2,text="Occupation: ",font="arial 13",bg=framebg).place(x=500,y=100)

mname=StringVar()
mname_entry=Entry(obj2,textvariable=mname,width=15,font="arial 10")
mname_entry.place(x=650,y=50)
mo=StringVar()
mo_entry=Entry(obj2,textvariable=mo,width=15,font="arial 10")
mo_entry.place(x=650,y=100)
#image
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="Student Registration System-images\Images/upload photo.png")
lbl=Label(f,bg='black',image=img)
lbl.place(x=0,y=0)


#button

Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=1000,y=370)
saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=Save)
saveButton.place(x=1000,y=450)
Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightpink",command=Clear).place(x=1000,y=530)
Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="grey",command=Exit).place(x=1000,y=610)


root.mainloop()

